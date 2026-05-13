import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _thin_border():
    thin = Side(style="thin", color="000000")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _make_cell_style(ws, row, col, value, bold=False, font_size=11, align_h="center", align_v="center", wrap=True, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Times New Roman", bold=bold, size=font_size, color="000000")
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)
    if border:
        cell.border = _thin_border()
    return cell

def build_ebs_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biểu mẫu EBS"

    # Row 1: PHỤ LỤC I:
    ws.merge_cells("A1:M1")
    cell1 = ws.cell(row=1, column=1, value="PHỤ LỤC I:")
    cell1.font = Font(name="Times New Roman", bold=True, size=12)
    cell1.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: BIỂU MẪU...
    ws.merge_cells("A2:M2")
    cell2 = ws.cell(row=2, column=1, value="BIỂU MẪU GHI NHẬN DẤU HIỆU CẢNH BÁO")
    cell2.font = Font(name="Times New Roman", bold=True, size=12)
    cell2.alignment = Alignment(horizontal="center", vertical="center")

    # Row 3: Đơn vị:
    ws.merge_cells("A3:M3")
    cell3 = ws.cell(row=3, column=1, value="Đơn vị: Hệ thống Epi Scout AI")
    cell3.font = Font(name="Times New Roman", bold=True, size=11)
    cell3.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # Headers setup
    headers_config = [
        # (StartCol, EndCol, StartRow, EndRow, Text)
        (1, 1, 4, 5, "Stt"),
        (2, 2, 4, 5, "Thời gian\nghi nhận\nthông tin"),
        (3, 7, 4, 4, "Thông tin về dấu hiệu cảnh báo"),
        (3, 3, 5, 5, "Nội dung"),
        (4, 4, 5, 5, "Nguồn\nthông\nbáo"),
        (5, 5, 5, 5, "Thời\ngian\nxảy\nra"),
        (6, 6, 5, 5, "Địa\nđiểm\nxảy\nra"),
        (7, 7, 5, 5, "Số mắc/\nchết/nhập\nviện hoặc\nkhả năng\nlây lan"),
        (8, 8, 4, 5, "Kết quả\nsàng\nlọc\n(xem\nhướng\ndẫn)"),
        (9, 9, 4, 5, "Kết quả\nxác\nminh\n(xem\nhướng\ndẫn)"),
        (10, 10, 4, 5, "Kết\nquả\nđánh\ngiá\nsự\nkiện"),
        (11, 11, 4, 5, "Thời\ngian\nbáo\ncáo\nlên\ntuyến\ntrên\n(nếu\ncó)"),
        (12, 12, 4, 5, "Các\nhoạt\nđộng\nđã\ntriển\nkhai\n(nếu\ncó)"),
        (13, 13, 4, 5, "Họ và\ntên\nngười\nghi\nnhận\nthông\ntin"),
    ]

    for start_col, end_col, start_row, end_row, text in headers_config:
        if start_col != end_col or start_row != end_row:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        
        # Apply border to all cells in the merged range
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = _thin_border()
                
        # Set value and style to top-left cell
        _make_cell_style(ws, start_row, start_col, text, bold=False, font_size=11)

    # Row 6: Column indices (0) to (12)
    for col in range(1, 14):
        _make_cell_style(ws, 6, col, f"({col-1})", bold=False, font_size=11, align_v="center")

    # Column widths
    col_widths = {
        1: 5,   # Stt
        2: 12,  # Thời gian ghi nhận
        3: 20,  # Nội dung
        4: 12,  # Nguồn
        5: 12,  # Thời gian xảy ra
        6: 15,  # Địa điểm
        7: 15,  # Số ca
        8: 12,  # KQ sàng lọc
        9: 12,  # KQ xác minh
        10: 10, # Đánh giá
        11: 12, # Báo cáo tuyến trên
        12: 15, # Đáp ứng
        13: 15  # Reporter
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
        
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 80
    ws.row_dimensions[6].height = 20
    
    wb.save('test_output.xlsx')
    print("Done!")

build_ebs_excel()
