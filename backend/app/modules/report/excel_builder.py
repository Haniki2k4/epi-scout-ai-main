"""
Excel Builder - Tạo file Excel theo Phụ lục I Quyết định 2018/QĐ-BYT
về Hướng dẫn Giám sát dựa vào Sự kiện (Event-Based Surveillance - EBS).

Cấu trúc bảng biểu mẫu:
STT | Ngày ghi nhận | Địa điểm (Thôn/Xã/Huyện/Tỉnh) | Mô tả dấu hiệu cảnh báo |
Nguồn tin | Kết quả sàng lọc | Kết quả xác minh | Đánh giá | Biện pháp đáp ứng
"""
import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter, column_index_from_string


# --- Style Constants ---
HEADER_BG_COLOR = "1565C0"   # Xanh dương đậm - màu Bộ Y tế
HEADER_FONT_COLOR = "FFFFFF" # Trắng
SUBHEADER_BG_COLOR = "E3F2FD" # Xanh nhạt cho dòng tiêu đề phụ
TABLE_HEADER_BG = "1976D2"   # Xanh tiêu đề bảng
TABLE_ALT_BG = "F5F9FF"      # Nền xen kẽ nhạt


def _thin_border() -> Border:
    thin = Side(style="thin", color="B0BEC5")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_cell_style(
    ws,
    row: int,
    col: int,
    value,
    bold: bool = False,
    font_size: int = 10,
    font_color: str = "212121",
    bg_color: Optional[str] = None,
    align_h: str = "left",
    align_v: str = "center",
    wrap: bool = True,
    border: bool = True,
):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(
        name="Times New Roman",
        bold=bold,
        size=font_size,
        color=font_color,
    )
    cell.alignment = Alignment(
        horizontal=align_h,
        vertical=align_v,
        wrap_text=wrap,
    )
    if bg_color:
        cell.fill = PatternFill(
            start_color=bg_color,
            end_color=bg_color,
            fill_type="solid",
        )
    if border:
        cell.border = _thin_border()
    return cell


def build_ebs_excel(report_data: dict, username: str = "Epi Scout AI") -> bytes:
    """
    Tạo file Excel Phụ lục I QĐ 2018/QĐ-BYT.
    
    Args:
        report_data: Dict trả về từ generator.get_report_data()
    
    Returns:
        bytes của file Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biểu mẫu EBS"

    # =========================================================
    # PHẦN TIÊU ĐỀ
    # =========================================================
    # Row 1: PHỤ LỤC I:
    ws.merge_cells("A1:M1")
    cell1 = ws.cell(row=1, column=1, value="PHỤ LỤC I:")
    cell1.font = Font(name="Times New Roman", bold=True, size=12, color="212121")
    cell1.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: BIỂU MẪU...
    ws.merge_cells("A2:M2")
    cell2 = ws.cell(row=2, column=1, value="BIỂU MẪU GHI NHẬN DẤU HIỆU CẢNH BÁO")
    cell2.font = Font(name="Times New Roman", bold=True, size=12, color="212121")
    cell2.alignment = Alignment(horizontal="center", vertical="center")

    # Thông tin đơn vị và ngày báo cáo
    generated_at: datetime = report_data.get("generated_at", datetime.utcnow())
    scope_hours: int = report_data.get("scope_hours", 72)
    start_date: datetime = report_data.get("start_date", generated_at)
    end_date: datetime = report_data.get("end_date", generated_at)

    # Row 3: Đơn vị:
    ws.merge_cells("A3:M3")
    cell3 = ws.cell(row=3, column=1, value="Đơn vị: Hệ thống Epi Scout AI")
    cell3.font = Font(name="Times New Roman", bold=True, size=11, color="212121")
    cell3.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 25

    # =========================================================
    # HEADER BẢNG (Row 4 & 5)
    # =========================================================
    headers_config = [
        # (StartCol, EndCol, StartRow, EndRow, Text)
        (1, 1, 4, 5, "Stt"),
        (2, 2, 4, 5, "Thời gian\nghi nhận\nthông tin"),
        (3, 7, 4, 4, "Thông tin về dấu hiệu cảnh báo"),
        (3, 3, 5, 5, "Nội dung"),
        (4, 4, 5, 5, "Nguồn\nthông\nbáo"),
        (5, 5, 5, 5, "Thời\ngian\nxảy\nra"),
        (6, 6, 5, 5, "Địa\nđiểm\nxảy\nra"),
        (7, 7, 5, 5, "Số mắc/\nchết/nhập\nviện hoặc\nkhả năng\nlây lan"),
        (8, 8, 4, 5, "Kết quả\nsàng\nlọc\n(xem\nhướng\ndẫn)"),
        (9, 9, 4, 5, "Kết quả\nxác\nminh\n(xem\nhướng\ndẫn)"),
        (10, 10, 4, 5, "Kết\nquả\nđánh\ngiá\nsự\nkiện"),
        (11, 11, 4, 5, "Thời\ngian\nbáo\ncáo\nlên\ntuyến\ntrên\n(nếu\ncó)"),
        (12, 12, 4, 5, "Các\nhoạt\nđộng\nđã\ntriển\nkhai\n(nếu\ncó)"),
        (13, 13, 4, 5, "Họ và\ntên\nngười\nghi\nnhận\nthông\ntin"),
    ]

    for start_col, end_col, start_row, end_row, text in headers_config:
        if start_col != end_col or start_row != end_row:
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = _thin_border()
                
        _make_cell_style(
            ws, start_row, start_col, text, 
            bold=False, font_size=11, align_h="center", align_v="center", wrap=True
        )

    # Row 6: Column indices (0) to (12)
    for col in range(1, 14):
        _make_cell_style(
            ws, 6, col, f"({col-1})", 
            bold=False, font_size=11, align_h="center", align_v="center", wrap=False
        )

    # Column widths
    col_widths = {
        1: 5,   # Stt
        2: 12,  # Thời gian ghi nhận
        3: 20,  # Nội dung
        4: 12,  # Nguồn
        5: 12,  # Thời gian xảy ra
        6: 15,  # Địa điểm
        7: 15,  # Số ca
        8: 12,  # KQ sàng lọc
        9: 12,  # KQ xác minh
        10: 10, # Đánh giá
        11: 12, # Báo cáo tuyến trên
        12: 15, # Đáp ứng
        13: 15  # Reporter
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
        
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 80
    ws.row_dimensions[6].height = 20

    # =========================================================
    # DỮ LIỆU (Rows 7+)
    # =========================================================
    top_events = report_data.get("top_events", [])
    alert_articles = report_data.get("alert_articles", [])
    recent_articles = report_data.get("recent_articles", [])

    rows_data = []

    # 1. Từ NewsEvent
    for event in top_events:
        location_str = event.location or ""
        desc = event.canonical_title
        case_info = f"Số ca: {event.case_count}" if event.case_count else "Chưa rõ số ca"

        rows_data.append({
            "date": event.event_date,
            "description": desc,
            "source": ", ".join(event.sources_preview[:3]) if event.sources_preview else "Hệ thống tự động",
            "time_happened": event.event_date,
            "location": location_str,
            "case_info": case_info,
            "screening": "",
            "verification": "",
            "assessment": event.severity or "Đang theo dõi",
            "time_reported": "",
            "response": "",
            "reporter": username
        })

    # 2. Bài báo có tag cảnh báo
    event_titles = {e.canonical_title for e in top_events}
    for article in alert_articles:
        if article.title and article.title not in event_titles:
            rows_data.append({
                "date": article.published_date,
                "description": article.title,
                "source": article.source or "Chưa xác định",
                "time_happened": article.published_date,
                "location": "",
                "case_info": "Chưa rõ",
                "screening": "",
                "verification": "",
                "assessment": "Cảnh báo",
                "time_reported": "",
                "response": "",
                "reporter": username
            })

    # 3. Bài báo gần nhất
    if len(rows_data) < 3:
        for article in recent_articles:
            if article.title and article.title not in event_titles:
                rows_data.append({
                    "date": article.published_date,
                    "description": article.title,
                    "source": article.source or "Chưa xác định",
                    "time_happened": article.published_date,
                    "location": "",
                    "case_info": "Chưa rõ",
                    "screening": "",
                    "verification": "",
                    "assessment": "Theo dõi",
                    "time_reported": "",
                    "response": "",
                    "reporter": username
                })
                if len(rows_data) >= 10:
                    break

    # Đảm bảo có ít nhất 5 dòng trống nếu không có dữ liệu
    if not rows_data:
        for _ in range(5):
            rows_data.append({k: "" for k in ["date", "description", "source", "time_happened", "location", "case_info", "screening", "verification", "assessment", "time_reported", "response", "reporter"]})

    data_start_row = 7
    for row_idx, row in enumerate(rows_data, start=data_start_row):
        is_alt = (row_idx % 2 == 0)
        row_bg = TABLE_ALT_BG if is_alt else None
        ws.row_dimensions[row_idx].height = 40

        date_val = row.get("date")
        date_str = date_val.strftime("%d/%m/%Y") if isinstance(date_val, datetime) else date_val

        time_happened_val = row.get("time_happened")
        time_happened_str = time_happened_val.strftime("%d/%m/%Y") if isinstance(time_happened_val, datetime) else time_happened_val

        values = [
            row_idx - data_start_row + 1,       # (0) STT
            date_str,                           # (1) Thời gian ghi nhận
            row.get("description", ""),         # (2) Nội dung
            row.get("source", ""),              # (3) Nguồn thông báo
            time_happened_str,                  # (4) Thời gian xảy ra
            row.get("location", ""),            # (5) Địa điểm xảy ra
            row.get("case_info", ""),           # (6) Số mắc/chết/nhập viện...
            row.get("screening", ""),           # (7) Kết quả sàng lọc
            row.get("verification", ""),        # (8) Kết quả xác minh
            row.get("assessment", ""),          # (9) Kết quả đánh giá
            row.get("time_reported", ""),       # (10) Thời gian báo cáo...
            row.get("response", ""),            # (11) Các hoạt động đã triển khai
            row.get("reporter", ""),            # (12) Họ và tên người ghi nhận
        ]
        
        for col_idx, val in enumerate(values, start=1):
            _make_cell_style(
                ws=ws,
                row=row_idx,
                col=col_idx,
                value=val,
                bold=False,
                font_size=11,
                bg_color=row_bg,
                align_h="center" if col_idx in (1, 2, 5, 8, 9, 10, 11) else "left",
                align_v="center",
                wrap=True,
            )

    # =========================================================
    # GHI CHÚ CUỐI
    # =========================================================
    note_row = data_start_row + len(rows_data) + 2
    ws.merge_cells(f"A{note_row}:M{note_row}")
    ws[f"A{note_row}"].value = (
        "* Ghi chú: Báo cáo được tạo tự động bởi Hệ thống Epi Scout AI dựa trên dữ liệu thu thập "
        f"trong {scope_hours} giờ (từ {start_date.strftime('%d/%m/%Y %H:%M')} đến {end_date.strftime('%d/%m/%Y %H:%M')}). "
        "Cán bộ y tế cần xác minh và bổ sung thông tin trước khi sử dụng chính thức."
    )
    ws[f"A{note_row}"].font = Font(name="Times New Roman", italic=True, size=10, color="616161")
    ws[f"A{note_row}"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 30

    # In ra bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
