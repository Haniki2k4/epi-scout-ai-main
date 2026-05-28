"""
Report Router - API endpoints cho module báo cáo dịch tễ.
"""
import io
import json
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ...core.database import get_db
from ..news.models import ArticleIdentity, ArticleDetails
from ..auth.security import get_current_active_user, require_admin_role
from .generator import get_report_data
from .docx_builder import build_word_report
from .excel_builder import build_ebs_excel
from .email_sender import send_report_email
from ...core.logger import get_logger

logger = get_logger("backend.report.router")

router = APIRouter(prefix="/api/report", tags=["report"])


# --- Schemas ---

class ReportRequest(BaseModel):
    scope_hours: int = 72  # Mặc định 72h = 3 ngày


class SendEmailRequest(BaseModel):
    scope_hours: int = 72
    attach_docx: bool = True
    attach_excel: bool = True
    custom_recipients: Optional[List[str]] = None  # Ghi đè danh sách người nhận


@router.post("/generate")
def generate_word_report(
    body: ReportRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """Tạo file báo cáo Word (.docx) từ dữ liệu trong DB."""
    logger.info("Word report requested | scope_hours={} user={}", body.scope_hours, current_user.username)

    try:
        report_data = get_report_data(db, scope_hours=body.scope_hours)
        docx_bytes = build_word_report(report_data)

        date_str = datetime.utcnow().strftime("%Y%m%d_%H%M")
        filename = f"BaoCao_DichBenh_{date_str}.docx"

        logger.info("Word report generated | size_bytes={}", len(docx_bytes))
        return Response(
            content=docx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error("Word report generation failed | error={}", str(e))
        raise HTTPException(status_code=500, detail=f"Tạo báo cáo thất bại: {str(e)}")


@router.post("/export-excel")
def export_ebs_excel(
    body: ReportRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """
    Xuất file Excel theo Phụ lục I Quyết định 2018/QĐ-BYT
    (Biểu mẫu Ghi nhận Dấu hiệu Cảnh báo EBS).
    """
    logger.info("Excel EBS report requested | scope_hours={} user={}", body.scope_hours, current_user.username)

    try:
        report_data = get_report_data(db, scope_hours=body.scope_hours)
        excel_bytes = build_ebs_excel(report_data, username=current_user.username)

        date_str = datetime.utcnow().strftime("%Y%m%d_%H%M")
        filename = f"BieuMau_EBS_QD2018_{date_str}.xlsx"

        logger.info("Excel EBS report generated | size_bytes={}", len(excel_bytes))
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        logger.error("Excel EBS report generation failed | error={}", str(e))
        raise HTTPException(status_code=500, detail=f"Xuất Excel thất bại: {str(e)}")


@router.post("/send-email")
def send_report_via_email(
    body: SendEmailRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user),
):
    """Tạo báo cáo và gửi qua SendGrid API đến danh sách email đã cấu hình."""
    logger.info(
        "Send email report requested | scope_hours={} attach_docx={} attach_excel={} user={}",
        body.scope_hours, body.attach_docx, body.attach_excel, current_user.username,
    )

    try:
        report_data = get_report_data(db, scope_hours=body.scope_hours)
        docx_bytes = build_word_report(report_data) if body.attach_docx else None
        excel_bytes = build_ebs_excel(report_data) if body.attach_excel else None

        recipients = body.custom_recipients
        if not recipients:
            from ..auth.models import User
            users = db.query(User).filter(
                User.is_active == True,
                User.email.isnot(None),
            ).all()
            recipients = [u.email for u in users if u.email]

        result = send_report_email(
            docx_bytes=docx_bytes,
            excel_bytes=excel_bytes,
            report_date=report_data["generated_at"],
            custom_recipients=recipients,
            report_data=report_data,
        )

        if not result["success"]:
            raise HTTPException(status_code=422, detail=result["message"])

        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Send email report failed | error={}", str(e))
        raise HTTPException(status_code=500, detail=f"Gửi email thất bại: {str(e)}")


# --- Hằng số ánh xạ cột Excel EBS (0-indexed) ---
# STT(0) | Ngày(1) | Địa điểm(2) | Mô tả(3) | Nguồn tin(4) |
# Kết quả sàng lọc(5) | Kết quả xác minh(6) | Đánh giá(7) | Biện pháp đáp ứng(8)
_COL_DESCRIPTION = 3   # "Mô tả dấu hiệu cảnh báo" — dùng để match tiêu đề bài báo
_COL_SCREENING = 5     # "Kết quả sàng lọc"
_COL_VERIFICATION = 6  # "Kết quả xác minh"
_COL_ASSESSMENT = 7    # "Đánh giá (Mức độ rủi ro)"
_COL_RESPONSE = 8      # "Biện pháp đáp ứng"

_EBS_MIN_COLS = 9       # Số cột tối thiểu của bảng EBS


def _safe_str(cell_value) -> str:
    """Chuyển giá trị ô về str, trả về '' nếu None."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _find_article_by_description(db: Session, description: str) -> ArticleIdentity | None:
    """
    Tìm bài báo khớp với mô tả EBS bằng cách so khớp tiêu đề (case-insensitive substring).
    Ưu tiên khớp chính xác, rồi đến khớp một phần (50 ký tự đầu).
    """
    if not description:
        return None

    # Thử tìm bài báo có tiêu đề chứa mô tả (hoặc ngược lại)
    # Giới hạn substring để tránh quét quá dài
    needle = description[:100].lower()

    # Tìm trong 2000 bài báo gần nhất (giới hạn phạm vi tìm kiếm)
    candidates = (
        db.query(ArticleIdentity)
        .order_by(ArticleIdentity.published_date.desc())
        .limit(2000)
        .all()
    )
    for article in candidates:
        title = (article.title or "").lower()
        if needle in title or title[:100] in description.lower():
            return article
    return None


@router.post("/import-excel")
async def import_labeled_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin=Depends(require_admin_role),
):
    """
    Nhập file Excel EBS đã dán nhãn thủ công.

    Cấu trúc bảng EBS được chấp nhận (9 cột):
    STT | Ngày | Địa điểm | Mô tả dấu hiệu | Nguồn tin |
    Kết quả sàng lọc | Kết quả xác minh | Đánh giá | Biện pháp đáp ứng

    Hệ thống sẽ đọc từng dòng dữ liệu (bỏ qua header/dòng tiêu đề),
    tìm bài báo tương ứng theo cột "Mô tả" rồi lưu nhãn thủ công.
    """
    # --- Kiểm tra định dạng ---
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400,
            detail="Chỉ chấp nhận file Excel (.xlsx hoặc .xls).",
        )

    logger.info("Import labeled Excel requested | filename={} user={}", filename, current_admin.username)

    try:
        import openpyxl

        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as exc:
        logger.error("Import Excel: failed to parse file | error={}", str(exc))
        raise HTTPException(status_code=422, detail=f"Không thể đọc file Excel: {exc}")

    # --- Tìm dòng header bảng dữ liệu ---
    # Dòng header EBS chứa từ "STT" ở cột đầu (thường là dòng 7 nhưng linh hoạt)
    data_start_row: int | None = None
    for row in ws.iter_rows():
        first_val = _safe_str(row[0].value).upper()
        if first_val == "STT":
            data_start_row = row[0].row + 1  # Dòng tiếp theo là dữ liệu
            break

    if data_start_row is None:
        # Không tìm thấy header → thử bắt đầu từ dòng 8 (mặc định EBS)
        data_start_row = 8

    # --- Parse từng dòng dữ liệu ---
    results = {"updated": 0, "skipped": 0, "not_found": 0, "errors": 0, "details": []}

    for row in ws.iter_rows(min_row=data_start_row):
        # Bỏ qua dòng có ít hơn số cột tối thiểu
        if len(row) < _EBS_MIN_COLS:
            results["skipped"] += 1
            continue

        # Bỏ qua dòng trống (STT trống)
        stt_val = _safe_str(row[0].value)
        if not stt_val or not stt_val.isdigit():
            results["skipped"] += 1
            continue

        description = _safe_str(row[_COL_DESCRIPTION].value)
        screening = _safe_str(row[_COL_SCREENING].value)
        verification = _safe_str(row[_COL_VERIFICATION].value)
        assessment = _safe_str(row[_COL_ASSESSMENT].value)
        response_text = _safe_str(row[_COL_RESPONSE].value)

        # Bỏ qua dòng không có nhãn thủ công nào
        has_label = any([screening, verification, assessment, response_text])
        if not has_label:
            results["skipped"] += 1
            continue

        # Tìm bài báo tương ứng
        try:
            article = _find_article_by_description(db, description)
        except Exception as exc:
            logger.warning("Import Excel: error finding article | desc={} error={}", description[:60], str(exc))
            results["errors"] += 1
            results["details"].append({"row": stt_val, "status": "error", "reason": str(exc)})
            continue

        if not article:
            results["not_found"] += 1
            results["details"].append({
                "row": stt_val,
                "status": "not_found",
                "description": description[:80],
            })
            continue

        # --- Cập nhật nhãn vào ArticleDetails.tags (JSON) ---
        try:
            details = article.details
            if not details:
                details = ArticleDetails(article_id=article.id)
                db.add(details)

            # Đọc tags hiện có (nếu là JSON dict), merge thêm nhãn EBS
            existing_tags: dict = {}
            if details.tags:
                try:
                    existing_tags = json.loads(details.tags)
                    if not isinstance(existing_tags, dict):
                        existing_tags = {"raw_tags": details.tags}
                except (json.JSONDecodeError, ValueError):
                    # Tags dạng plain text cũ → giữ lại
                    existing_tags = {"raw_tags": details.tags}

            # Ghi nhãn EBS, chỉ ghi đè khi có giá trị mới
            if screening:
                existing_tags["ebs_screening"] = screening
            if verification:
                existing_tags["ebs_verification"] = verification
            if assessment:
                existing_tags["ebs_assessment"] = assessment
            if response_text:
                existing_tags["ebs_response"] = response_text

            details.tags = json.dumps(existing_tags, ensure_ascii=False)
            db.commit()

            results["updated"] += 1
            results["details"].append({
                "row": stt_val,
                "status": "updated",
                "article_id": article.id,
                "title": (article.title or "")[:80],
            })
        except Exception as exc:
            db.rollback()
            logger.error("Import Excel: failed to update article | id={} error={}", article.id, str(exc))
            results["errors"] += 1
            results["details"].append({
                "row": stt_val,
                "status": "error",
                "article_id": article.id,
                "reason": str(exc),
            })

    logger.info(
        "Import Excel completed | updated={} skipped={} not_found={} errors={}",
        results["updated"], results["skipped"], results["not_found"], results["errors"],
    )
    return {
        "status": "ok",
        "filename": filename,
        "summary": {
            "updated": results["updated"],
            "skipped": results["skipped"],
            "not_found": results["not_found"],
            "errors": results["errors"],
        },
        "details": results["details"][:50],  # Giới hạn 50 dòng chi tiết trả về
    }
