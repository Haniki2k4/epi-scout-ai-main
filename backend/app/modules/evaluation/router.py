import io
import os
import re
import unicodedata

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from pydantic import BaseModel
from ...core.database import get_db
from ..auth import security
from ..news.models import ArticleIdentity
from . import crud, schemas, models

router = APIRouter(prefix="/api/evaluation", tags=["Evaluation"])

VALID_EVALUATION_LABELS = {"relevant", "noise", "irrelevant", "unsure"}
EVALUATION_DATASET_FILENAME = "llm_evaluation_dataset.xlsx"


# --- Schema riêng cho endpoint evaluation/articles ---

class DiseaseCaseInfo(BaseModel):
    disease_name: str
    case_count: int
    location: Optional[str] = None

class ArticleForEvaluation(BaseModel):
    id: int
    title: Optional[str] = None
    link: str
    source: Optional[str] = None
    keywords_matched: Optional[str] = None
    event_id: Optional[int] = None
    llm_label: str          # "relevant" | "irrelevant"
    human_label: Optional[str] = None   # Nhãn thực từ ArticleEvaluation
    keyword_is_correct: Optional[bool] = None
    corrected_keyword: Optional[str] = None
    is_verified: bool = False
    cases: List[DiseaseCaseInfo] = []

    class Config:
        from_attributes = True


class EvaluationArticlesResponse(BaseModel):
    items: List[ArticleForEvaluation]
    total: int
    total_verified: int


def _normalize_excel_key(value) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", ascii_text).strip()


def _normalize_label(value) -> str | None:
    label = _normalize_excel_key(value)
    if label in {"relevant", "lien quan"}:
        return "relevant"
    if label in {"noise", "nhieu", "co de cap nhung khong co su kien"}:
        return "noise"
    if label in {"irrelevant", "khong lien quan", "khong phu hop"}:
        return "irrelevant"
    if label in {"unsure", "khong chac", "khong ro", "chua ro", "can xem lai"}:
        return "unsure"
    return None


def _find_import_columns(sheet) -> tuple[int, dict[str, int]] | None:
    aliases = {
        "article_id": {"id", "article id", "ma bai bao"},
        "title": {"title", "headline", "tieu de"},
        "summary": {"summary", "description", "sapo", "tom tat", "mo ta"},
        "link": {"link", "url"},
        "llm_label": {"llm label", "nhan llm", "du doan", "label llm"},
        "human_label": {"human label", "nhan thu cong", "nhan human", "ground truth", "label dung"},
    }

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        normalized_headers = [_normalize_excel_key(cell) for cell in row]
        columns: dict[str, int] = {}

        for field, field_aliases in aliases.items():
            for index, header in enumerate(normalized_headers):
                if any(alias == header or alias in header for alias in field_aliases):
                    columns[field] = index
                    break

        has_article_key = any(key in columns for key in ("article_id", "link", "title"))
        if has_article_key and "human_label" in columns:
            return row_index, columns

    return None


def _get_cell(row: tuple, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def _safe_text(value) -> str:
    return str(value or "").strip()


def _find_article_for_import(db: Session, row: tuple, columns: dict[str, int]) -> ArticleIdentity | None:
    raw_id = _get_cell(row, columns.get("article_id"))
    if raw_id is not None and str(raw_id).strip():
        try:
            article_id = int(float(str(raw_id).strip()))
            article = db.query(ArticleIdentity).filter(ArticleIdentity.id == article_id).first()
            if article:
                return article
        except ValueError:
            pass

    link = _safe_text(_get_cell(row, columns.get("link")))
    if link:
        article = db.query(ArticleIdentity).filter(ArticleIdentity.link == link).first()
        if article:
            return article

    title = _safe_text(_get_cell(row, columns.get("title")))
    if title:
        return db.query(ArticleIdentity).filter(ArticleIdentity.title == title).first()

    return None


def _dataset_output_path() -> str:
    configured_path = os.getenv("LLM_FEW_SHOT_DATASET_PATH", "").strip()
    if configured_path and os.path.isabs(configured_path):
        return configured_path

    backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
    filename = configured_path or EVALUATION_DATASET_FILENAME
    if os.path.dirname(filename):
        return os.path.abspath(os.path.join(backend_dir, filename))
    return os.path.join(backend_dir, "data", filename)


def _write_labeled_dataset_for_llm(rows: list[dict[str, str]]) -> str | None:
    if not rows:
        return None

    import openpyxl

    output_path = _dataset_output_path()
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Evaluation Data"
    sheet.append(["title", "summary", "llm_label", "human_label"])
    for row in rows:
        sheet.append([row["title"], row["summary"], row["llm_label"], row["human_label"]])
    workbook.save(output_path)
    return output_path

@router.get("/metrics")
def get_metrics(db: Session = Depends(get_db)):
    return crud.get_metrics(db)


@router.get("/articles", response_model=EvaluationArticlesResponse)
def get_articles_for_evaluation(
    skip: int = 0,
    limit: int = 100,
    filter_label: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin=Depends(security.require_admin_role),
):
    """
    Lấy danh sách bài báo dùng để gán nhãn thủ công.

    - Trả về `cases[]` (tên dịch bệnh từ bảng disease_cases).
    - Trả về `llm_label` thực từ ArticleEvaluation (nếu có),
      hoặc suy ra từ event_id (relevant/irrelevant).
    - Trả về `human_label` và `is_verified` từ ArticleEvaluation.
    """
    from ..news.models import DiseaseCase

    query = db.query(ArticleIdentity)
    if filter_label:
        query = query.join(models.ArticleEvaluation).filter(models.ArticleEvaluation.human_label == filter_label)

    total = query.count()
    total_verified = db.query(models.ArticleEvaluation).join(ArticleIdentity).filter(
        models.ArticleEvaluation.human_label.isnot(None)
    ).count()

    articles = (
        query
        .options(
            joinedload(ArticleIdentity.details),
            joinedload(ArticleIdentity.cases),
        )
        .order_by(ArticleIdentity.published_date.desc(), ArticleIdentity.id.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    # Batch-load evaluations cho tất cả bài viết trong 1 query (thay vì N+1)
    article_ids = [a.id for a in articles]
    eval_map: dict[int, models.ArticleEvaluation] = {}
    if article_ids:
        evals = db.query(models.ArticleEvaluation).filter(
            models.ArticleEvaluation.article_id.in_(article_ids)
        ).all()
        eval_map = {e.article_id: e for e in evals}

    results: List[ArticleForEvaluation] = []
    for a in articles:
        eval_rec = eval_map.get(a.id)

        # llm_label: ưu tiên lấy từ DB, fallback về event_id
        if eval_rec and eval_rec.llm_label:
            llm_label = eval_rec.llm_label
        else:
            llm_label = "relevant" if a.event_id else "irrelevant"

        cases_info = [
            DiseaseCaseInfo(
                disease_name=c.disease_name,
                case_count=c.case_count,
                location=c.location,
            )
            for c in (a.cases or [])
        ]

        results.append(
            ArticleForEvaluation(
                id=a.id,
                title=a.title,
                link=a.link,
                source=a.details.source if a.details else None,
                keywords_matched=a.details.keywords_matched if a.details else None,
                event_id=a.event_id,
                llm_label=llm_label,
                human_label=eval_rec.human_label if eval_rec else None,
                keyword_is_correct=eval_rec.keyword_is_correct if eval_rec else None,
                corrected_keyword=eval_rec.corrected_keyword if eval_rec else None,
                is_verified=eval_rec.is_verified if eval_rec else False,
                cases=cases_info,
            )
        )

    return {
        "items": results,
        "total": total,
        "total_verified": total_verified,
    }

@router.put("/{article_id}", response_model=schemas.ArticleEvaluationDTO)
def update_human_label(
    article_id: int,
    body: schemas.ArticleEvaluationUpdate,
    db: Session = Depends(get_db),
    current_admin=Depends(security.require_admin_role)
):
    eval_record = crud.update_human_label(
        db, 
        article_id, 
        body.human_label, 
        current_admin.id,
        keyword_is_correct=body.keyword_is_correct,
        corrected_keyword=body.corrected_keyword,
        update_article_keyword=body.update_article_keyword
    )
    return eval_record


@router.post("/import-excel")
async def import_evaluations_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin=Depends(security.require_admin_role),
):
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(status_code=400, detail="Chỉ chấp nhận file Excel (.xlsx, .xlsm hoặc .xls).")

    try:
        import openpyxl

        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Không thể đọc file Excel: {exc}")

    column_match = _find_import_columns(sheet)
    if not column_match:
        workbook.close()
        raise HTTPException(
            status_code=422,
            detail="Không tìm thấy cột ID/Link/Tiêu đề và cột Nhãn Thủ công trong file Excel.",
        )

    header_row, columns = column_match
    from datetime import datetime

    summary = {"updated": 0, "skipped": 0, "not_found": 0, "errors": 0, "dataset_examples": 0}
    details: list[dict[str, str | int]] = []
    dataset_rows: list[dict[str, str]] = []

    # 1. Thu thập dữ liệu từ file Excel
    raw_rows = []
    all_article_ids = set()
    all_links = set()
    all_titles = set()
    
    for excel_row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(row):
            summary["skipped"] += 1
            continue

        human_label = _normalize_label(_get_cell(row, columns.get("human_label")))
        if human_label not in VALID_EVALUATION_LABELS:
            summary["skipped"] += 1
            continue
            
        raw_id = _get_cell(row, columns.get("article_id"))
        if raw_id is not None and str(raw_id).strip():
            try:
                all_article_ids.add(int(float(str(raw_id).strip())))
            except ValueError:
                pass
                
        link = _safe_text(_get_cell(row, columns.get("link")))
        if link:
            all_links.add(link)
            
        title = _safe_text(_get_cell(row, columns.get("title")))
        if title:
            all_titles.add(title)
            
        raw_rows.append((excel_row_number, row, human_label))

    # 2. Batch query các bài viết
    articles_by_id = {a.id: a for a in db.query(ArticleIdentity).filter(ArticleIdentity.id.in_(all_article_ids)).all()} if all_article_ids else {}
    articles_by_link = {a.link: a for a in db.query(ArticleIdentity).filter(ArticleIdentity.link.in_(all_links)).all()} if all_links else {}
    articles_by_title = {a.title: a for a in db.query(ArticleIdentity).filter(ArticleIdentity.title.in_(all_titles)).all()} if all_titles else {}

    def find_article_in_memory(r):
        r_id = _get_cell(r, columns.get("article_id"))
        if r_id is not None and str(r_id).strip():
            try:
                aid = int(float(str(r_id).strip()))
                if aid in articles_by_id: return articles_by_id[aid]
            except ValueError:
                pass
        lnk = _safe_text(_get_cell(r, columns.get("link")))
        if lnk and lnk in articles_by_link: return articles_by_link[lnk]
        ttl = _safe_text(_get_cell(r, columns.get("title")))
        if ttl and ttl in articles_by_title: return articles_by_title[ttl]
        return None

    # 3. Batch query evaluations
    found_articles = set()
    article_mapping = {}
    for excel_row_number, row, human_label in raw_rows:
        article = find_article_in_memory(row)
        if article:
            found_articles.add(article.id)
            article_mapping[excel_row_number] = article
            
    eval_map = {}
    if found_articles:
        evals = db.query(models.ArticleEvaluation).filter(models.ArticleEvaluation.article_id.in_(found_articles)).all()
        eval_map = {e.article_id: e for e in evals}
        
    # 4. Xử lý gán nhãn hàng loạt
    for excel_row_number, row, human_label in raw_rows:
        try:
            article = article_mapping.get(excel_row_number)
            if not article:
                summary["not_found"] += 1
                details.append({"row": excel_row_number, "status": "not_found"})
                continue

            imported_llm_label = _normalize_label(_get_cell(row, columns.get("llm_label")))
            fallback_llm_label = "relevant" if article.event_id else "irrelevant"
            llm_label = imported_llm_label or fallback_llm_label
            
            # Cập nhật trực tiếp không qua crud.update_human_label để tránh db.commit() nhiều lần
            eval_record = eval_map.get(article.id)
            if not eval_record:
                eval_record = models.ArticleEvaluation(article_id=article.id)
                db.add(eval_record)
                eval_map[article.id] = eval_record
                
            eval_record.llm_label = llm_label
            eval_record.human_label = human_label
            eval_record.is_verified = True
            eval_record.verified_at = datetime.utcnow()
            eval_record.verified_by = current_admin.id

            title = _safe_text(article.title) or _safe_text(_get_cell(row, columns.get("title")))
            summary_text = _safe_text(article.summary) or _safe_text(_get_cell(row, columns.get("summary")))
            if title and summary_text:
                dataset_rows.append({
                    "title": title,
                    "summary": summary_text,
                    "llm_label": eval_record.llm_label,
                    "human_label": human_label,
                })

            summary["updated"] += 1
            details.append({
                "row": excel_row_number,
                "status": "updated",
                "article_id": article.id,
                "title": title[:80],
            })
        except Exception as exc:
            summary["errors"] += 1
            details.append({"row": excel_row_number, "status": "error", "reason": str(exc)})

    # Lưu toàn bộ thay đổi 1 lần
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi khi lưu dữ liệu vào database: {str(e)}")

    workbook.close()

    summary["dataset_examples"] = len(dataset_rows)
    dataset_path = None
    dataset_error = None
    try:
        dataset_path = _write_labeled_dataset_for_llm(dataset_rows)
    except Exception as exc:
        dataset_error = str(exc)
        summary["errors"] += 1

    return {
        "status": "ok",
        "filename": filename,
        "summary": summary,
        "dataset_path": dataset_path,
        "dataset_error": dataset_error,
        "details": details[:50],
    }

@router.get("/export-excel")
def export_evaluations_excel(
    db: Session = Depends(get_db),
    current_admin=Depends(security.require_admin_role)
):
    import openpyxl
    from fastapi.responses import StreamingResponse
    import io

    # Lấy các bài báo mới nhất
    # Để đơn giản, lấy 500 bài báo gần đây
    articles = db.query(ArticleIdentity).order_by(ArticleIdentity.published_date.desc()).limit(500).all()
    
    # Batch-load evaluations cho 500 bài viết trong 1 query (tránh N+1)
    article_ids = [a.id for a in articles]
    eval_map = {}
    if article_ids:
        evals = db.query(models.ArticleEvaluation).filter(
            models.ArticleEvaluation.article_id.in_(article_ids)
        ).all()
        eval_map = {e.article_id: e for e in evals}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Data"
    
    # Thêm 4 nhãn giải thích
    ws.append(["Giải thích nhãn:"])
    ws.append(["- relevant: Có sự kiện dịch tễ thực"])
    ws.append(["- noise: Đề cập bệnh nhưng là tư vấn/dịch vụ/thảo luận"])
    ws.append(["- irrelevant: Không liên quan đến dịch bệnh"])
    ws.append(["- unsure: Chưa chắc, cần xem xét thêm"])
    ws.append([]) 
    
    headers = ["ID", "Tiêu đề", "Tóm tắt", "Link", "Nguồn", "Keyword Dịch bệnh", "Nhãn LLM", "Nhãn Thủ công"]
    ws.append(headers)
    header_row = ws.max_row
    
    for a in articles:
        # Lấy từ dict thay vì query DB cho từng bài (N+1)
        eval_rec = eval_map.get(a.id)
        llm_lbl = eval_rec.llm_label if eval_rec and eval_rec.llm_label else ("relevant" if a.event_id else "irrelevant")
        human_lbl = eval_rec.human_label if eval_rec and eval_rec.human_label else ""
        
        ws.append([
            a.id,
            a.title,
            a.summary or "",
            a.link,
            a.details.source if a.details else "",
            a.details.keywords_matched if a.details else "",
            llm_lbl,
            human_lbl
        ])
        
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"relevant,noise,irrelevant,unsure"', allow_blank=True)
    ws.add_data_validation(dv)
    if ws.max_row > header_row:
        dv.add(f"H{header_row + 1}:H{ws.max_row}")
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=llm_evaluation_dataset.xlsx"}
    )

@router.post("/sync-dataset")
def sync_dataset_to_file(db: Session = Depends(get_db), current_admin=Depends(security.require_admin_role)):
    evals = db.query(models.ArticleEvaluation).filter(models.ArticleEvaluation.is_verified == True).all()
    dataset_rows = []
    orphaned_deleted = False
    
    for e in evals:
        if not e.article:
            db.delete(e)
            orphaned_deleted = True
            continue
            
        if e.human_label in VALID_EVALUATION_LABELS:
            title = _safe_text(e.article.title)
            summary_text = _safe_text(e.article.summary)
            if title and summary_text:
                dataset_rows.append({
                    "title": title,
                    "summary": summary_text,
                    "llm_label": e.llm_label or "relevant",
                    "human_label": e.human_label,
                })
                
    if orphaned_deleted:
        db.commit()
    
    if not dataset_rows:
        raise HTTPException(status_code=400, detail="Không có dữ liệu gán nhãn hợp lệ để đồng bộ.")
        
    try:
        dataset_path = _write_labeled_dataset_for_llm(dataset_rows)
        return {"message": f"Đã đồng bộ {len(dataset_rows)} mẫu dữ liệu cho mô hình.", "path": dataset_path}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Lỗi ghi file: {str(exc)}")

@router.post("/exclude-irrelevant")
def exclude_irrelevant_articles(db: Session = Depends(get_db), current_admin=Depends(security.require_admin_role)):
    """Đánh dấu các bài báo noise/irrelevant là is_excluded để ẩn khỏi hiển thị công khai."""
    evals = db.query(models.ArticleEvaluation).filter(
        models.ArticleEvaluation.human_label.in_(["noise", "irrelevant"])
    ).all()
    
    article_ids = [e.article_id for e in evals]
    if not article_ids:
        return {"message": "Không có bài báo nào cần ẩn", "excluded_count": 0}
        
    db.query(ArticleIdentity).filter(ArticleIdentity.id.in_(article_ids)).update(
        {ArticleIdentity.is_excluded: True}, synchronize_session=False
    )
    
    db.commit()
    return {"message": f"Đã ẩn {len(article_ids)} bài báo không relevant.", "excluded_count": len(article_ids)}
